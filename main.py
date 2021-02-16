from openpyxl import load_workbook
import subprocess
import sys
from PyQt5 import QtWidgets, uic
from PyQt5.QtWidgets import QMainWindow
import shutil
import os

# Словарь дней недели:
# ------------------------------------------------------------------------------------------
day = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24,
       25, 26, 27, 28, 29, 30, 31]
letters = ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V',
           'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ']
day_letters = dict(zip(day, letters))

# -------------------------------------------------------------------------------------------


class MainWindow(QMainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        uic.loadUi('ui.ui', self)
        for n in range(14, 22):
            getattr(self, 'Plus_%s' % n).pressed.connect(self.costs)
        self.Plus_22.pressed.connect(self.reserve)
        self.Plus_23.pressed.connect(self.salary)
        self.lineEdit_22.returnPressed.connect(self.Plus_22.click)
        self.lineEdit_23.returnPressed.connect(self.Plus_23.click)
        # self.file_btn.pressed.connect(self.get_file)
        # self.get_file()
        self.openBookBtn.pressed.connect(self.open_book_btn)
        self.createBtn.pressed.connect(self.cr_table)
        self.readBtn.pressed.connect(self.read_costs)
        for i in range(14, 22):
            getattr(self, 'lineEdit_%s' % i).returnPressed.connect(getattr(self, 'Plus_%s' % i).click)

    # Создать таблицу
    def cr_table(self):
        date = self.calendarWidget.selectedDate()  # Дата из календаря
        date_month = str(date.month())
        date_year = str(date.year())
        file_name = date_year + '_' + date_month + '.xlsx'
        # Если файла с таким именем нет, то копируем из шаблона
        if not os.path.exists("d:/_google disk/Финансы/" + file_name):

            shutil.copyfile(r"d:/_google disk/Финансы/template.xlsx",
                            r"d:/_google disk/Финансы/" + file_name)

            # Записываем заголовок таблицы
            book = load_workbook("d:/_google disk/Финансы/" + file_name)
            sheet = book.active
            sheet['A1'] = 'Финансы' + ' ' + date.toString('MMMM, yyyy')
            book.save("d:/_google disk/Финансы/" + file_name)

        # Если такого файла нет, показать ошибку
        else:
            msg = QtWidgets.QMessageBox()
            msg.setText("Такая таблица уже есть!\n\nВыберите другой месяц!")
            msg.setIcon(QtWidgets.QMessageBox.Critical)
            msg.setWindowTitle('ОШИБКА!')
            msg.setStyleSheet("color : white ;"
                              "background-color: #2f3237; ")
            msg.exec_()

    # Резерв
    def reserve(self):
        date = self.calendarWidget.selectedDate()  # Дата из календаря
        file_name = str(date.year()) + '_' + str(date.month()) + '.xlsx'  # Имя файла
        file = "d:/_google disk/Финансы/" + file_name  # Путь к файлу
        try:
            book = load_workbook(file)
            sheet = book.active
            text = self.lineEdit_22.text()  # Значение поля ввода
            if text != '':
                try:
                    if sheet['M5'].value is None:
                        sheet['M5'] = '=' + text
                        book.save(file)
                        self.lineEdit_22.setText('')
                    # Если поле уже заполнено прибавляем значение к текущему
                    else:
                        sheet['M5'] = sheet['M5'].value + '+' + text
                        book.save(file)
                        self.lineEdit_22.setText('')
                except PermissionError:
                    self.error()
            else:
                print('empty')
        except FileNotFoundError:
            self.error_no_file()

    # Зарплата
    def salary(self):
        date = self.calendarWidget.selectedDate()  # Дата из календаря
        file_name = str(date.year()) + '_' + str(date.month()) + '.xlsx'  # Имя файла
        file = "d:/_google disk/Финансы/" + file_name  # Путь к файлу
        try:
            book = load_workbook(file)  # Открыть файл xl
            sheet = book.active  # Выбрать активный лист

            text = self.lineEdit_23.text()  # Значение поля ввода

            # Если значение ячейки пустое, то просто вписываем текст из поля ввода
            if text != '':
                try:
                    if sheet['E5'].value is None:
                        sheet['E5'] = '=' + text
                        book.save(file)
                        self.lineEdit_23.setText('')
                    # Если поле уже заполнено прибавляем значение к текущему
                    else:
                        sheet['E5'] = sheet['E5'].value + '+' + text
                        book.save(file)
                        self.lineEdit_23.setText('')
                        print(sheet['E5'].value)
                except PermissionError:
                    self.error()
            else:
                print('empty')

        except FileNotFoundError:
            self.error_no_file()

    # Расходы
    def costs(self):
        date = self.calendarWidget.selectedDate()  # Дата из календаря
        file_name = str(date.year()) + '_' + str(date.month()) + '.xlsx'  # Имя файла
        file = "d:/_google disk/Финансы/" + file_name  # Путь к файлу
        try:
            book = load_workbook(file)  # Открыть файл excel
            sheet = book.active  # Выбрать активный лист
            n_day = self.calendarWidget.selectedDate().day()  # Номер дня недели
            letter = day_letters[n_day]  # Выбор ячейки по столбцу
            line = str(self.sender().objectName().split('_')[-1])  # Из имени объекта кнопки берем только число
            cell = letter + line  # Готовая  ячейка
            text = getattr(self, 'lineEdit_%s' % line).text()  # Значение поля ввода

            # Если значение ячейки пустое, то просто вписываем текст из поля ввода
            if text != '':
                try:
                    if sheet[cell].value is None:
                        sheet[cell] = '=' + text
                        book.save(file)
                        getattr(self, 'lineEdit_%s' % line).setText('')
                    # Если поле уже заполнено прибавляем значение к текущему
                    else:
                        sheet[cell] = sheet[cell].value + '+' + text
                        book.save(file)
                        getattr(self, 'lineEdit_%s' % line).setText('')
                except PermissionError:
                    self.error()
            else:
                print('empty')

        except FileNotFoundError:
            self.error_no_file()

    # Прочитать ячейки
    def read_costs(self):
        try:
            date = self.calendarWidget.selectedDate()  # Дата из календаря
            file_name = str(date.year()) + '_' + str(date.month()) + '.xlsx'  # Имя файла
            file = "d:/_google disk/Финансы/" + file_name  # Путь к файлу

            book = load_workbook(file, data_only=True)  # Открыть файл excel
            sheet = book.active  # Выбрать активный лист
            n_day = self.calendarWidget.selectedDate().day()  # Номер дня недели
            letter = day_letters[n_day]  # Выбор ячейки по столбцу

            for line in range(14,22):
                cell = letter + str(line)
                if sheet[cell].value is not None:
                    getattr(self, 'lineEdit_%s' % line).setText(str(sheet[cell].value))
                else:
                    getattr(self, 'lineEdit_%s' % line).setText('')
        except FileNotFoundError:
            self.error_no_file()
        except PermissionError:
            self.error()

    # Функция показа ошибки при открытом файле
    def error(self):
        msg = QtWidgets.QMessageBox()
        msg.setText("Закройте файл!")
        msg.setIcon(QtWidgets.QMessageBox.Critical)
        msg.setWindowTitle('ОШИБКА!')
        msg.setStyleSheet("color : white ;"
                          "background-color: #2f3237; ")
        msg.exec_()

    # Ошибка при отсутсвии файла в дериктории
    def error_no_file(self):
        msg = QtWidgets.QMessageBox()
        msg.setText("Таблицы не существует!")
        msg.setIcon(QtWidgets.QMessageBox.Critical)
        msg.setWindowTitle('ОШИБКА!')
        msg.setStyleSheet("color : white ;"
                          "background-color: #2f3237; ")
        msg.exec_()

    # Функция для кнопки открытия файла
    def open_book_btn(self):
        date = self.calendarWidget.selectedDate()  # Дата из календаря
        file_name = str(date.year()) + '_' + str(date.month()) + '.xlsx'  # Имя файла
        file = "d:/_google disk/Финансы/" + file_name  # Путь к файлу
        if os.path.exists(file):
            subprocess.call(file, shell=True)
        else:
            self.error_no_file()


def main():
    app = QtWidgets.QApplication(sys.argv)
    app.setStyle('Fusion')
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
